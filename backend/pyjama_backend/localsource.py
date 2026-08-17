"""Local file ingestion → encrypted managed source (Phase 5 — Epic U, Sources refactor).

Reads a user-provided CSV / XLSX / Parquet file into an Arrow table, then stores
an encrypted managed copy as a standalone *source* (sources.py) — reusable across
any workspace, independent of which workspace (if any) imported it. The user's
original file is left untouched (it is user-owned, §9.4).
"""

from __future__ import annotations

import io

import pyarrow as pa
import pyarrow.parquet as pq

from . import crypto
from .keystore import KeyStore
from .sources import SourceManifest, source_data_dir, write_manifest


class LocalSourceError(Exception):
    pass


def detect_format(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".parquet"):
        return "parquet"
    if lower.endswith(".xlsx"):
        return "xlsx"
    raise LocalSourceError(f"unsupported file type: {filename}")


def _read_csv(data: bytes) -> pa.Table:
    import pyarrow.csv as pacsv

    return pacsv.read_csv(pa.BufferReader(data))


def _read_xlsx(data: bytes) -> pa.Table:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise LocalSourceError("empty spreadsheet")
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    body = rows[1:]
    cols = {h: [r[i] if i < len(r) else None for r in body] for i, h in enumerate(header)}
    return pa.table(cols)


def _read(fmt: str, data: bytes) -> pa.Table:
    if fmt == "parquet":
        return pq.read_table(io.BytesIO(data))
    if fmt == "csv":
        return _read_csv(data)
    if fmt == "xlsx":
        return _read_xlsx(data)
    raise LocalSourceError(f"unsupported format: {fmt}")


def inspect(fmt: str, data: bytes) -> dict:
    """Peek at a file's schema/row count without importing it."""
    table = _read(fmt, data)
    return {"columns": list(table.column_names), "row_count": table.num_rows}


def import_bytes(keystore: KeyStore, name: str, fmt: str, data: bytes, local_path: str | None = None) -> SourceManifest:
    """Import a file as a standalone encrypted source."""
    table = _read(fmt, data)
    return import_arrow_table(keystore, name, fmt, table, local_path=local_path)


def import_arrow_table(keystore: KeyStore, name: str, kind: str, table: pa.Table, local_path: str | None = None) -> SourceManifest:
    """Store an already-built Arrow table as a new standalone source."""
    from datetime import datetime, timezone

    from .sources import create_placeholder

    m = create_placeholder(name, kind)
    wdek = crypto.load_or_create_wdek(keystore, m.source_id)
    dest_dir = source_data_dir(m.source_id)
    dest_dir.mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    pq.write_table(table, buf, compression="zstd")
    (dest_dir / "source-00000.parquet").write_bytes(crypto.encrypt(wdek, buf.getvalue()))

    m.columns = list(table.column_names)
    m.row_count = table.num_rows
    m.logical_bytes = buf.tell()
    m.partition_files = 1
    m.encryption_key_id = crypto.wdek_key_name(m.source_id)
    m.local_path = local_path
    m.refreshed_at = datetime.now(timezone.utc).isoformat()
    write_manifest(m)
    return m


def refresh_from_path(keystore: KeyStore, source_id: str) -> SourceManifest:
    """Re-read a local file source from its original path (P9's source-refresh)."""
    from datetime import datetime, timezone

    from . import sources as sources_mod

    m = sources_mod.read_manifest(source_id)
    if not m.local_path:
        raise LocalSourceError("this source has no original file path to refresh from")
    from pathlib import Path

    p = Path(m.local_path)
    if not p.is_file():
        raise LocalSourceError(f"original file no longer found: {m.local_path}")
    fmt = detect_format(p.name)
    table = _read(fmt, p.read_bytes())

    sources_mod.clear_data(source_id)
    wdek = crypto.get_wdek(keystore, source_id)
    buf = io.BytesIO()
    pq.write_table(table, buf, compression="zstd")
    (sources_mod.source_data_dir(source_id) / "source-00000.parquet").write_bytes(crypto.encrypt(wdek, buf.getvalue()))

    m.columns = list(table.column_names)
    m.row_count = table.num_rows
    m.logical_bytes = buf.tell()
    m.partition_files = 1
    m.refreshed_at = datetime.now(timezone.utc).isoformat()
    sources_mod.write_manifest(m)
    return m
